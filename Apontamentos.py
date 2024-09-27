import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, timedelta

def apply_borders(ws):
    # Define o estilo da borda
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Aplica a borda a todas as células do worksheet
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border

# Função para gerar todas as datas no intervalo definido
def generate_date_range(start_date, end_date):
    return pd.date_range(start=start_date, end=end_date)

# Função para remover o dia da semana da data no JSON
def clean_json_date(date_str):
    # Remove o dia da semana (últimos 3 caracteres)
    return date_str[:10]


# Função para fazer a requisição e salvar os dados filtrados no Excel
def get_filtered_data_and_save_to_excel(url, payload, headers, output_path):
    response = requests.post(url, json=payload, headers=headers)

    if response.status_code == 200:
        try:
            data = response.json()
            if "Obj" in data and isinstance(data['Obj'], list):
                all_entries = []
                all_fixed_data = []

                # Defina o intervalo de datas do relatório
                start_date = datetime.strptime(payload["DataInicio"], "%d/%m/%Y")
                end_date = datetime.strptime(payload["DataFim"], "%d/%m/%Y")
                full_date_range = generate_date_range(start_date, end_date)

                # Itera sobre cada objeto no campo "Obj"
                for item in data['Obj']:
                    # Adiciona os dados fixos para cada funcionário
                    fixed_data = {
                        'Empresa': item['InfoEmpresa']['Nome'],
                        'CNPJ': item['InfoEmpresa']['CNPJCPF'],
                        'PIS': item['InfoFuncionario']['PIS'],
                        'Funcionario': item['InfoFuncionario']['Nome'],
                        'Matricula': item['InfoFuncionario']['Matricula']
                    }

                    # Cria um dicionário para armazenar as marcações por data
                    entradas_por_data = {}
                    for entrada in item['Entradas']:
                        # Limpa a data removendo o dia da semana
                        data_limpa = clean_json_date(entrada['Data'])
                        # Converte a data para o formato datetime
                        entrada_data_formatada = datetime.strptime(data_limpa, "%d/%m/%Y")
                        entradas_por_data[entrada_data_formatada] = entrada

                    # Para cada data no intervalo, insere os dados ou deixa em branco
                    for date in full_date_range:
                        if date in entradas_por_data:
                            entrada = entradas_por_data[date]
                            entry_data = {
                                'Data': date.strftime("%d/%m/%Y"),
                                'Horario': entrada['Horario'],
                                'Apontamentos': entrada['Apontamentos'],
                                'Horas Trabalhadas': entrada['HTrab'],
                                'Horas Extras': entrada['HE'],
                                'Descontos': entrada['Descontos'],
                                'Debito': entrada['Debito'],
                                'Credito': entrada['Credito']
                            }
                        else:
                            # Preenche com valores em branco para datas sem entrada
                            entry_data = {
                                'Data': date.strftime("%d/%m/%Y"),
                                'Horario': '',
                                'Apontamentos': '',
                                'Horas Trabalhadas': '',
                                'Horas Extras': '',
                                'Descontos': '',
                                'Debito': '',
                                'Credito': ''
                            }

                        # Adiciona os dados fixos e de entrada combinados
                        combined_data = {**fixed_data, **entry_data}
                        all_entries.append(combined_data)

                # Converte a lista de dados em um DataFrame
                final_df = pd.DataFrame(all_entries)

                # Adiciona a nova coluna "Justificativas" com valores em branco
                final_df['Justificativas'] = ''
                final_df['Entrada'] = ''
                final_df['Saida Pausa'] = ''
                final_df['Volta Pausa'] = ''
                final_df['Saida'] = ''



                # Cria um Workbook do openpyxl
                wb = Workbook()
                ws = wb.active
                
                justificativa_options = ['Atestado', 'Folga', 'Justificativa de Horas', 'Abono']
                justificativa_str = ','.join(justificativa_options)
                
                justificativa_validation = DataValidation(type='list', formula1=f'"{justificativa_str}"', allow_blank=True)
                justificativa_validation.error = 'Escolher valores da lista'
                justificativa_validation.errorTitle = 'Entrada Invalida'
                justificativa_validation.prompt = 'Selecione uma justificativa'
                justificativa_validation.promptTitle = 'Justificativas'
                
                justificativa_col_index = final_df.columns.get_loc("Justificativas") + 1
                justificativa_col_letter = ws.cell(row=1, column=justificativa_col_index).column_letter
                
                ws.add_data_validation(justificativa_validation)
                justificativa_validation.add(f'{justificativa_col_letter}2:{justificativa_col_letter}{len(final_df)+1}')

                # Escreve o DataFrame no Excel usando openpyxl
                for r_idx, row in enumerate(dataframe_to_rows(final_df, index=False, header=True), 1):
                    for c_idx, value in enumerate(row, 1):
                        ws.cell(row=r_idx, column=c_idx, value=value)

                # Ajusta a largura das colunas automaticamente
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column].width = adjusted_width

                # Formatação de cor alternada nas linhas
                fill_grey = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                
                

                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    for cell in row:
                        # Alterna entre as cores
                        if cell.row % 2 == 0:
                            cell.fill = fill_grey
                        else:
                            cell.fill = fill_white
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                        
                apply_borders(ws)

                # Salva o arquivo Excel com formatação
                wb.save(output_path)
                print(f"Arquivo Excel salvo com sucesso em {output_path}")
            else:
                print("Nenhum dado no campo 'Obj'.")
        except ValueError as e:
            print(f"Erro ao decodificar JSON: {e}")
            print("Conteúdo da resposta:")
            print(response.text)
    else:
        print(f"Falha: {response.status_code}")
        print(response.text)
        
        

# URL do endpoint e payload
url = 'https://www.dimepkairos.com.br/RestServiceApi/ReportEmployeePunch/GetReportEmployeePunch'  # Modifique para o endpoint correto
payload = {"param": "value"}  # Modifique conforme necessário

# Headers da requisição
headers = {
        "identifier": "31.487.442/0001-38",
        "key": "a06a8bc3-9b7f-4d97-b45d-15549eee8063",
        'User-Agent': 'PostmanRuntime/7.30.0'
}

payload = {
        "MatriculaPessoa": [], # Manter em branco para coletar todos os colaboradores
        "DataInicio":"", # Inserir data de inicio da coleta
        "DataFim":"", # Inserir data final da coleta
        "ResponseType":"AS400V1" # Campo Fixo não alterar
}

# Local onde o arquivo gerado será salvo.
output_path = r'C:\Users\luis.marques\Desktop\VsCode\Casa.xlsx'

# Chama a função para obter e filtrar os dados e salvar no Excel
get_filtered_data_and_save_to_excel(url, payload, headers, output_path)
